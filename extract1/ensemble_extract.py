"""
双模型集成人名提取
- 模型A (finetune_best.pt) + 模型B (frozen_best.pt) 分别提取
- 一致 → 直接输出
- 不一致 → LLM 裁定
用法: python ensemble_extract.py <input.xlsx> [output.xlsx]
"""
import sys
import torch
import openpyxl
import re

from config import BIO_ID2LABEL, BIO_LABELS
from model import ErnieCRF, ErnieCRF2
from config import ERNIE_LOCAL, CHECKPOINT, CHECKPOINT_FROZEN


def clean_names(names: list) -> list:
    cleaned = []
    for n in names:
        n = re.sub(r"等\d*人?$", "", n)
        if n.strip():
            cleaned.append(n.strip())
    return cleaned


def expand_bracket_names(names: list, title: str) -> list:
    expanded = []
    for name in names:
        idx = title.find(name)
        if idx == -1:
            expanded.append(name)
            continue
        found = False
        # 英文（中文）
        if idx > 0 and title[idx - 1] == "（":
            right = idx + len(name)
            if right < len(title) and title[right] == "）":
                left = idx - 2
                if left >= 0 and re.match(r"[A-Za-z]", title[left]):
                    start = left
                    while start > 0 and re.match(r"[A-Za-z]", title[start - 1]):
                        start -= 1
                    found = True
                    expanded.append(title[start:right + 1])
        # 中文（英文）
        if not found:
            right = idx + len(name)
            if right < len(title) and title[right] == "（":
                close = title.find("）", right)
                if close != -1 and re.match(r"[A-Za-z]", title[right + 1:close]):
                    found = True
                    expanded.append(title[idx:close + 1])
        if not found:
            expanded.append(name)
    # 去子串
    deduped = []
    for n in expanded:
        if not any(n != other and n in other for other in expanded):
            deduped.append(n)
    return deduped


def postprocess(names, title):
    names = clean_names(names)
    names = expand_bracket_names(names, title)
    names = [n for n in names if n in title]
    return names


def model_extract(title, model, tokenizer, device):
    chars = list(title)
    ids = [tokenizer.cls_token_id]
    for c in chars:
        ids.extend(tokenizer.encode(c, add_special_tokens=False))
    ids.append(tokenizer.sep_token_id)

    input_ids = torch.tensor([ids], device=device)
    mask = torch.ones_like(input_ids)
    with torch.no_grad():
        preds = model(input_ids, mask)[0]
    preds = preds[1:1 + len(chars)]

    names, cur = [], []
    for char, lid in zip(title, preds):
        tag = BIO_ID2LABEL.get(lid, "O")
        if tag == "B-PER":
            if cur:
                names.append("".join(cur))
            cur = [char]
        elif tag == "I-PER" and cur:
            cur.append(char)
        else:
            if cur:
                names.append("".join(cur))
                cur = []
    if cur:
        names.append("".join(cur))
    return names


def load_models(device):
    from transformers import AutoTokenizer
    tokenizer = AutoTokenizer.from_pretrained(ERNIE_LOCAL)

    model_a = ErnieCRF(ERNIE_LOCAL, len(BIO_LABELS)).to(device)
    model_a.load_state_dict(torch.load(CHECKPOINT, map_location=device, weights_only=True))
    model_a.eval()

    model_b = ErnieCRF(ERNIE_LOCAL, len(BIO_LABELS)).to(device)
    model_b.load_state_dict(torch.load(CHECKPOINT_FROZEN, map_location=device, weights_only=True))
    model_b.eval()

    return tokenizer, model_a, model_b


def main():
    if len(sys.argv) < 2:
        print("用法: python ensemble_extract.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_xlsx = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "提取结果_集成.xlsx"

    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"加载双模型 ({device})...")
    tokenizer, model_a, model_b = load_models(device)

    print(f"读取: {input_xlsx}")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["A列：原数据", "B列：提取人名", "C列：方法", "D列：姓名字数"])

    stats = {"一致": 0, "LLM裁定": 0, "未检出": 0}
    disagreements = []  # (row_idx, title, names_a, names_b)
    results_cache = {}  # title → final_names

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=12, values_only=True)):
        title = str(row[0]) if row[0] else ""
        org = str(row[11]) if len(row) > 11 and row[11] else ""

        if not title:
            continue

        clean_title = title
        if org and org in title:
            clean_title = title.replace(org, "").replace("  ", " ").strip()

        # 双模型分别提取
        raw_a = model_extract(clean_title, model_a, tokenizer, device)
        raw_b = model_extract(clean_title, model_b, tokenizer, device)
        names_a = postprocess(raw_a, title)
        names_b = postprocess(raw_b, title)

        name_str_a = "、".join(names_a) if names_a else ""
        name_str_b = "、".join(names_b) if names_b else ""

        if name_str_a == name_str_b:
            # 一致，直接输出
            stats["一致"] += 1
            if not name_str_a:
                stats["未检出"] += 1
            name_len = "、".join(str(len(n)) for n in names_a) if names_a else "0"
            out_ws.append([title, name_str_a, "一致", name_len])
        else:
            # 不一致，缓存等 LLM
            disagreements.append((title, name_str_a, name_str_b))

    print(f"一致: {stats['一致']} 条 (含未检出)")
    print(f"不一致需LLM裁定: {len(disagreements)} 条")

    if disagreements:
        from llm_resolver import resolve_batch
        print(f"\n开始LLM裁定...")
        final_names = resolve_batch(disagreements)

        for (title, na, nb), final in zip(disagreements, final_names):
            stats["LLM裁定"] += 1
            names = [n.strip() for n in final.split("、") if n.strip()] if final else []
            # 后处理
            names = [n for n in names if n in title]
            name_str = "、".join(names) if names else ""
            name_len = "、".join(str(len(n)) for n in names) if names else "0"
            out_ws.append([title, name_str, "LLM裁定", name_len])

    out_wb.save(output_xlsx)

    print(f"\n=== 统计 ===")
    print(f"一致:     {stats['一致']} 条")
    print(f"LLM裁定: {stats['LLM裁定']} 条")
    print(f"未检出:   {stats['未检出']} 条")
    print(f"已保存: {output_xlsx}")


if __name__ == "__main__":
    main()
