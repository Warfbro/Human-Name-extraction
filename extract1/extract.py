"""
人名提取程序 — ERNIE+CRF 模型
用法: python extract.py <input.xlsx> [output.xlsx]
"""
import sys
import torch
import openpyxl

import re

from config import BIO_ID2LABEL
from model import load_model


def clean_names(names: list) -> list:
    """后处理：去掉 '等N人'、'等N' 后缀"""
    cleaned = []
    for n in names:
        n = re.sub(r"等\d*人?$", "", n)
        if n.strip():
            cleaned.append(n.strip())
    return cleaned


def expand_bracket_names(names: list, title: str) -> list:
    """如果名字旁有括号且括号旁是英文，扩展为 英文+括号+中文名 整体"""
    expanded = []
    for name in names:
        idx = title.find(name)
        if idx == -1:
            expanded.append(name)
            continue

        found = False

        # 模式1: EnglishName（中文名） — 名字在括号内
        if idx > 0 and title[idx - 1] == "（":
            # 找右边的 ）
            right = idx + len(name)
            if right < len(title) and title[right] == "）":
                # 找左边的英文
                left = idx - 2
                if left >= 0 and re.match(r"[A-Za-z]", title[left]):
                    start = left
                    while start > 0 and re.match(r"[A-Za-z]", title[start - 1]):
                        start -= 1
                    found = True
                    expanded.append(title[start:right + 1])

        # 模式2: （中文名）EnglishName — 名字在括号内，英文在右边
        if not found and idx > 0 and title[idx - 1] == "（":
            right = idx + len(name)
            if right < len(title) and title[right] == "）":
                after = right + 1
                if after < len(title) and re.match(r"[A-Za-z]", title[after]):
                    end = after
                    while end + 1 < len(title) and re.match(r"[A-Za-z]", title[end + 1]):
                        end += 1
                    found = True
                    expanded.append(title[idx - 1:end + 1])

        # 模式3: 中文名（EnglishName） — 名字在括号左边
        if not found:
            right = idx + len(name)
            if right < len(title) and title[right] == "（":
                # 找右边的 ）
                close = title.find("）", right)
                if close != -1:
                    between = title[right + 1:close]
                    if re.match(r"[A-Za-z]", between):
                        found = True
                        expanded.append(title[idx:close + 1])

        # 模式4: ）EnglishName（中文名 — 名字左右都有括号+英文
        # 已经被上面覆盖，不额外处理

        if not found:
            expanded.append(name)

    # 去重：若某名字是另一名字的子串，去掉短的
    deduped = []
    for n in expanded:
        if not any(n != other and n in other for other in expanded):
            deduped.append(n)
    return deduped


def model_extract(title, model, tokenizer, device):
    """用模型从标题中提取人名"""
    chars = list(title)
    ids = [tokenizer.cls_token_id]
    for c in chars:
        ids.extend(tokenizer.encode(c, add_special_tokens=False))
    ids.append(tokenizer.sep_token_id)

    input_ids = torch.tensor([ids], device=device)
    mask = torch.ones_like(input_ids)
    with torch.no_grad():
        preds = model(input_ids, mask)[0]
    preds = preds[1:1 + len(chars)]

    names, cur = [], []
    for char, lid in zip(title, preds):
        tag = BIO_ID2LABEL.get(lid, "O")
        if tag == "B-PER":
            if cur:
                names.append("".join(cur))
            cur = [char]
        elif tag == "I-PER" and cur:
            cur.append(char)
        else:
            if cur:
                names.append("".join(cur))
                cur = []
    if cur:
        names.append("".join(cur))
    return names


def main():
    if len(sys.argv) < 2:
        print("用法: python extract.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_xlsx = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "提取结果.xlsx"
    use_frozen = "--frozen" in sys.argv or "--fc2" not in sys.argv  # 默认 frozen
    use_fc2 = "--fc2" in sys.argv

    # 加载模型
    device = "cuda" if torch.cuda.is_available() else "cpu"
    if use_fc2:
        model_name = "fc2"
    elif use_frozen:
        model_name = "frozen"
    else:
        model_name = "finetune"
    print(f"加载模型 ({device}, {model_name})...")
    model, tokenizer = load_model(device, frozen=use_frozen, fc2=use_fc2)

    # 读取
    print(f"读取: {input_xlsx}")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active

    # 输出 xlsx
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["A列：原数据", "B列：提取人名", "C列：方法", "D列：姓名字数"])

    stats = {"模型": 0, "未检出": 0}

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=12, values_only=True):
        title = str(row[0]) if row[0] else ""
        org = str(row[11]) if len(row) > 11 and row[11] else ""  # L列=被许可对象

        if not title:
            continue

        # 预处理: 用L列机构名清洗标题
        clean_title = title
        if org and org in title:
            clean_title = title.replace(org, "").replace("  ", " ").strip()

        names = model_extract(clean_title, model, tokenizer, device)
        method = "模型" if names else "未检出"
        stats[method] += 1

        names = clean_names(names)
        names = expand_bracket_names(names, title)
        # 后处理: 检查提取人名是否在原标题中存在，不存在则丢弃
        names = [n for n in names if n in title]
        name_str = "、".join(names) if names else ""
        name_len = "、".join(str(len(n)) for n in names) if names else "0"
        out_ws.append([title, name_str, method, name_len])

    out_wb.save(output_xlsx)

    print(f"模型提取: {stats['模型']} 条")
    print(f"未检出:   {stats['未检出']} 条")
    print(f"已保存: {output_xlsx}")


if __name__ == "__main__":
    main()
